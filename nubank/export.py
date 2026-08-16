"""Estagio 5: o unico estagio que toca disco.

Escreve na aba Cartao da planilha de planejamento, com backup timestampado
antes de qualquer escrita e dry-run por padrao.

Sobre as colunas B e H, que e a parte que se erra facil:

    B  'Fatura total'      = total a pagar, o que sai do bolso no vencimento
    H  'Gasto do periodo'  = compras + IOF + outros lancamentos

Sao grandezas diferentes, e a diferenca e exatamente `fatura anterior -
pagamentos`: o saldo que veio da fatura passada. No mes em que voce paga R$ 50
a mais, B fica R$ 50 abaixo de H.

Quem decompoe C+D+E e H, nao B. Por isso a formula da coluna F ('Nao
classificado') confere contra H, e nao contra B - so assim ela zera quando
esta tudo classificado, que e para o que ela existe. Conferir contra B faria F
mostrar o saldo carregado, um numero grande e vermelho numa coluna cujo nome
promete outra coisa.

Como a planilha nasceu sem a coluna H, o export cuida do layout tambem:
escreve o cabecalho de H e reescreve as formulas de F. E idempotente.
"""

from __future__ import annotations

import csv
import shutil
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

from .errors import ErroExport
from .models import Bucket, Fatura
from .money import format_brl, to_float
from .xlsx import Planilha

ABA = "Cartao"
LINHA_CABECALHO = 3
COLUNA_MES = "A"

# Coluna da planilha <- bucket. F fica de fora: e formula, escrita pelo layout.
COLUNAS = {
    "B": None,  # Fatura total: total a pagar, nao soma de bucket
    "C": Bucket.PARCELAS,
    "D": Bucket.ESSENCIAL,
    "E": Bucket.SUPERFLUO,
}

COLUNA_GASTO = "H"
CABECALHO_GASTO = "Gasto do periodo"
LARGURA_GASTO = 18

# Os tres baldes que C, D e E decompoem.
_BUCKETS_PLANILHA = (Bucket.PARCELAS, Bucket.ESSENCIAL, Bucket.SUPERFLUO)

# Colunas de onde o estilo de H e copiado, para H nascer com o mesmo formato de
# moeda e o mesmo visual de cabecalho do resto da aba.
MOLDE_CABECALHO = "E"
MOLDE_VALOR = "E"

CABECALHO_ESPERADO = {
    "A": "Mes",
    "B": "Fatura total",
    "C": "Parcelas e assinaturas",
}


def formula_nao_classificado(linha: int) -> str:
    """F confere contra H (gasto do periodo), nao contra B (total a pagar)."""
    return f'=IF($B{linha}="","",${COLUNA_GASTO}{linha}-SUM($C{linha}:$E{linha}))'


def formula_total(coluna: str, primeira: int, ultima: int) -> str:
    return f"=SUM({coluna}{primeira}:{coluna}{ultima})"


@dataclass(frozen=True)
class Mudanca:
    celula: str
    rotulo: str
    antes: object
    depois: object

    @property
    def mudou(self) -> bool:
        if isinstance(self.antes, float) and isinstance(self.depois, float):
            return abs(self.antes - self.depois) > 1e-9
        return self.antes != self.depois

    def _fmt(self, valor: object) -> str:
        if valor is None or valor == "":
            return "(vazio)"
        if isinstance(valor, float):
            return format_brl(Decimal(str(valor)))
        return str(valor)

    def __str__(self) -> str:
        return f"{self.celula} {self.rotulo:<24} {self._fmt(self.antes)} -> {self._fmt(self.depois)}"


def gasto_do_periodo(fatura: Fatura) -> Decimal:
    """O que a coluna H recebe: a soma dos lancamentos do periodo.

    Sai da soma das transacoes, e nao de `compras + IOF + outros` do resumo do
    PDF, ainda que o resumo pareca a fonte mais natural. Os dois divergem em 1
    centavo em metade das faturas, por arredondamento do proprio Nubank, e como
    C/D/E vem das transacoes, usar o resumo faria a coluna F ficar oscilando
    entre 0,01 e -0,01 - ruido numa coluna cujo unico trabalho e mostrar zero
    quando esta tudo classificado.

    Aqui entram tambem os lancamentos ainda sem categoria. E o que faz a coluna
    F mostrar exatamente quanto falta classificar quando o export roda com
    --ignorar-pendentes.
    """
    return sum((t.valor for t in fatura.transacoes if t.eh_gasto), Decimal("0.00"))


def observacao(fatura: Fatura) -> str:
    """Texto da coluna G.

    O saldo carregado da fatura anterior nao tem coluna propria - e a diferenca
    entre B e H. Fica escrito aqui para a diferenca entre os dois nao virar
    caca ao bug.
    """
    periodo = f"{fatura.periodo_ini:%d/%m}-{fatura.periodo_fim:%d/%m}"
    n = sum(1 for t in fatura.transacoes if t.eh_gasto)
    texto = f"{periodo} | {n} lanc."
    # Calculado sobre B e H para bater com o que a planilha mostra, em vez de
    # sobre fatura_anterior - pagamentos, que difere pelo centavo do resumo.
    saldo = fatura.total_a_pagar - gasto_do_periodo(fatura)
    if saldo:
        texto += (
            f" | B-H = {format_brl(saldo)}: saldo da fatura anterior"
            f" ({format_brl(fatura.fatura_anterior)} devidos,"
            f" {format_brl(fatura.pagamentos)} pagos)"
        )
    return texto


def _linhas_de_dados(ws) -> list[int]:
    """As linhas de mes: coluna A com data. Exclui cabecalho e a linha TOTAL."""
    return [
        linha
        for linha in range(LINHA_CABECALHO + 1, ws.max_row + 1)
        if isinstance(ws[f"{COLUNA_MES}{linha}"].value, datetime)
    ]


def _localiza_total(ws) -> int | None:
    for linha in range(LINHA_CABECALHO + 1, ws.max_row + 1):
        valor = ws[f"{COLUNA_MES}{linha}"].value
        if isinstance(valor, str) and valor.strip().upper().startswith("TOTAL"):
            return linha
    return None


def _mudancas_layout(ws) -> list[Mudanca]:
    """O que falta na aba para a coluna H existir e a F conferir contra ela.

    Roda em todo export e e idempotente: na segunda vez nada muda.
    """
    mudancas: list[Mudanca] = []
    celula_cab = f"{COLUNA_GASTO}{LINHA_CABECALHO}"
    mudancas.append(
        Mudanca(
            celula=f"layout   {celula_cab}",
            rotulo="cabecalho de H",
            antes=ws[celula_cab].value,
            depois=CABECALHO_GASTO,
        )
    )

    linhas = _linhas_de_dados(ws)
    for linha in linhas:
        celula = f"F{linha}"
        mudancas.append(
            Mudanca(
                celula=f"layout   {celula}",
                rotulo="formula nao classif.",
                antes=ws[celula].value,
                depois=formula_nao_classificado(linha),
            )
        )

    total = _localiza_total(ws)
    if total and linhas:
        celula = f"{COLUNA_GASTO}{total}"
        mudancas.append(
            Mudanca(
                celula=f"layout   {celula}",
                rotulo="total de H",
                antes=ws[celula].value,
                depois=formula_total(COLUNA_GASTO, linhas[0], linhas[-1]),
            )
        )
    return mudancas


def _alvos(fatura: Fatura) -> dict[str, object]:
    """Coluna -> valor que a fatura deve gravar. Uma fonte so, usada pelo
    dry-run e pela escrita, para os dois nunca discordarem."""
    return {
        "B": to_float(fatura.total_a_pagar),
        "C": to_float(fatura.total_bucket(Bucket.PARCELAS)),
        "D": to_float(fatura.total_bucket(Bucket.ESSENCIAL)),
        "E": to_float(fatura.total_bucket(Bucket.SUPERFLUO)),
        "G": observacao(fatura),
        COLUNA_GASTO: to_float(gasto_do_periodo(fatura)),
    }


def _localiza_linha(ws, competencia: str) -> int:
    ano, mes = (int(p) for p in competencia.split("-"))
    for linha in range(LINHA_CABECALHO + 1, ws.max_row + 1):
        valor = ws[f"{COLUNA_MES}{linha}"].value
        if isinstance(valor, datetime) and valor.year == ano and valor.month == mes:
            return linha
    raise ErroExport(
        f"a aba '{ABA}' nao tem linha para a competencia {competencia}. "
        f"A coluna {COLUNA_MES} precisa ter uma data desse mes."
    )


def _valida_planilha(ws) -> None:
    for coluna, esperado in CABECALHO_ESPERADO.items():
        atual = ws[f"{coluna}{LINHA_CABECALHO}"].value
        if atual is None or not str(atual).strip().lower().startswith(esperado.lower()):
            raise ErroExport(
                f"a aba '{ABA}' nao tem o formato esperado: {coluna}{LINHA_CABECALHO} "
                f"deveria comecar com '{esperado}', mas contem {atual!r}."
            )


def planeja(planilha: str | Path, faturas: list[Fatura]) -> list[Mudanca]:
    """Calcula as mudancas sem escrever nada."""
    planilha = Path(planilha)
    if not planilha.exists():
        raise ErroExport(f"planilha nao encontrada: {planilha}")

    wb = openpyxl.load_workbook(planilha)
    if ABA not in wb.sheetnames:
        raise ErroExport(
            f"a planilha nao tem a aba '{ABA}'. Abas: {', '.join(wb.sheetnames)}"
        )
    ws = wb[ABA]
    _valida_planilha(ws)

    mudancas: list[Mudanca] = _mudancas_layout(ws)
    for fatura in faturas:
        linha = _localiza_linha(ws, fatura.competencia)
        for coluna, novo in _alvos(fatura).items():
            celula = f"{coluna}{linha}"
            # H so ganha cabecalho na etapa de layout, que ainda nao rodou aqui.
            padrao = CABECALHO_GASTO if coluna == COLUNA_GASTO else coluna
            rotulo = str(ws[f"{coluna}{LINHA_CABECALHO}"].value or padrao)
            mudancas.append(
                Mudanca(
                    celula=f"{fatura.competencia} {celula}",
                    rotulo=rotulo[:24],
                    antes=ws[celula].value,
                    depois=novo,
                )
            )
    wb.close()
    return mudancas


def faz_backup(planilha: Path, pasta: Path | None = None) -> Path:
    pasta = pasta or planilha.parent / "backups"
    pasta.mkdir(parents=True, exist_ok=True)
    carimbo = datetime.now().strftime("%Y%m%d-%H%M%S")
    destino = pasta / f"{planilha.stem}-{carimbo}{planilha.suffix}"
    shutil.copy2(planilha, destino)
    return destino


def aplica(planilha: str | Path, faturas: list[Fatura]) -> tuple[Path, list[Mudanca]]:
    """Escreve de verdade. Devolve (caminho do backup, mudancas aplicadas).

    A escrita e cirurgica (ver nubank/xlsx.py): so o XML da aba Cartao e o
    calcPr do workbook sao tocados, e todas as outras partes do zip saem byte a
    byte iguais. openpyxl e usado apenas para localizar as linhas, o que e
    leitura e nao danifica nada.
    """
    planilha = Path(planilha)
    mudancas = [m for m in planeja(planilha, faturas) if m.mudou]

    wb = openpyxl.load_workbook(planilha)
    ws = wb[ABA]
    linhas_por_competencia = {
        f.competencia: _localiza_linha(ws, f.competencia) for f in faturas
    }
    linhas_dados = _linhas_de_dados(ws)
    linha_total = _localiza_total(ws)
    wb.close()

    backup = faz_backup(planilha)

    doc = Planilha(planilha)
    aba = doc.aba(ABA)

    # Layout primeiro: H precisa existir com o estilo da vizinha antes de
    # receber valor, senao entra sem formato de moeda.
    aba.define_texto(
        f"{COLUNA_GASTO}{LINHA_CABECALHO}",
        CABECALHO_GASTO,
        estilo=aba.estilo(f"{MOLDE_CABECALHO}{LINHA_CABECALHO}"),
    )
    aba.largura_coluna(COLUNA_GASTO, LARGURA_GASTO)

    # O valor de F por competencia: o que sobrou sem classificar, zero quando
    # esta tudo. Vai como cache junto da formula, para quem abre o arquivo sem
    # avaliar formula (preview do editor, visualizador leve) ver o numero em
    # vez de tentar calcular e mostrar erro.
    residuo = {
        linhas_por_competencia[f.competencia]: to_float(
            gasto_do_periodo(f)
            - sum(
                (f.total_bucket(b) for b in _BUCKETS_PLANILHA),
                start=Decimal("0.00"),
            )
        )
        for f in faturas
    }

    # F4 e a mestra de uma formula compartilhada (ref F4:F15): reescrever o
    # texto dela ja desloca as irmas por linha. As demais so entram aqui se a
    # planilha nao usar formula compartilhada.
    for linha in linhas_dados:
        celula = f"F{linha}"
        if aba.formula(celula) is not None:
            aba.define_formula(
                celula,
                formula_nao_classificado(linha),
                valor=residuo.get(linha),
            )

    for fatura in faturas:
        linha = linhas_por_competencia[fatura.competencia]
        molde = aba.estilo(f"{MOLDE_VALOR}{linha}")
        for coluna, valor in _alvos(fatura).items():
            ref = f"{coluna}{linha}"
            if isinstance(valor, str):
                aba.define_texto(ref, valor)
            else:
                estilo = None if aba.existe(ref) else molde
                aba.define_numero(ref, valor, estilo=estilo)

    # Depois dos valores de H, para o total poder somar o que foi escrito.
    if linha_total and linhas_dados:
        aba.define_formula(
            f"{COLUNA_GASTO}{linha_total}",
            formula_total(COLUNA_GASTO, linhas_dados[0], linhas_dados[-1]),
            estilo=aba.estilo(f"{MOLDE_VALOR}{linha_total}"),
            valor=to_float(
                sum(
                    (gasto_do_periodo(f) for f in faturas), start=Decimal("0.00")
                )
            ),
        )

    doc.recalcular_ao_abrir()
    doc.salva()
    return backup, mudancas


def exporta_csv(destino: str | Path, faturas: list[Fatura]) -> Path:
    """CSV de detalhe, uma linha por transacao. Para conferir a mao quando quiser."""
    destino = Path(destino)
    destino.parent.mkdir(parents=True, exist_ok=True)
    with open(destino, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(
            [
                "competencia", "data", "descricao", "merchant", "valor", "tipo",
                "categoria", "bucket", "parcela", "cartao", "moeda_origem",
                "valor_origem", "origem_categoria",
            ]
        )
        for fatura in faturas:
            for t in fatura.transacoes:
                parcela = (
                    f"{t.parcela_atual}/{t.parcela_total}" if t.parcela_total else ""
                )
                w.writerow(
                    [
                        fatura.competencia,
                        t.data.isoformat(),
                        t.descricao_raw,
                        t.merchant_norm,
                        str(t.valor).replace(".", ","),
                        t.tipo.value,
                        t.categoria or "",
                        t.bucket.value if t.bucket else "",
                        parcela,
                        t.cartao_final or "",
                        t.moeda_origem or "",
                        str(t.valor_origem).replace(".", ",") if t.valor_origem else "",
                        t.origem_categoria.value if t.origem_categoria else "",
                    ]
                )
    return destino
