"""Confere as formulas que ligam as abas da planilha, e repara as quebradas.

Existe por um motivo observado, nao hipotetico: seis formulas da coluna
'Gastos fixos' do Controle Mensal foram substituidas pelo texto 'Aporte mensal
planejado' quando a planilha foi salva por um aplicativo externo. Na vez
seguinte ja eram nove.

O estrago e silencioso. A celula fica com um texto no lugar da formula, o Excel
nao acusa erro nenhum, e a conta segue: 'Total saidas' encolhe, 'Saldo do mes'
cresce, e o Painel mostra uma folga que voce nao tem. Nada fica vermelho.

So sao conferidas as formulas que atravessam abas - as que o usuario nao edita
e que, se sumirem, quebram o encadeamento inteiro. Formula que voce pode
legitimamente querer mudar fica de fora.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from .export import (
    ABA,
    COLUNA_GASTO,
    formula_nao_classificado,
    formula_total,
)
from .xlsx import Planilha

ABA_CONTROLE = "Controle Mensal"
ABA_GASTOS = "Gastos Fixos"

PRIMEIRA_LINHA = 4
ULTIMA_LINHA = 15
LINHA_TOTAL_GASTOS = 13


@dataclass(frozen=True)
class Regra:
    """Uma formula que precisa estar viva para o encadeamento funcionar."""

    aba: str
    celula: str
    formula: str  # sem o '=' inicial, como o OOXML guarda
    o_que: str


@dataclass(frozen=True)
class Problema:
    regra: Regra
    encontrado: str | None

    @property
    def sumiu(self) -> bool:
        return self.encontrado is None

    def __str__(self) -> str:
        atual = "(sem formula)" if self.sumiu else self.encontrado
        return (
            f"{self.regra.aba}!{self.regra.celula}  {self.regra.o_que}\n"
            f"      esperado: {self.regra.formula}\n"
            f"      achado:   {atual}"
        )


def regras() -> list[Regra]:
    """As ligacoes entre abas que a planilha depende."""
    lista: list[Regra] = []

    for i, linha in enumerate(range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1), start=1):
        lista.append(
            Regra(
                aba=ABA_CONTROLE,
                celula=f"E{linha}",
                formula=(
                    f"INDEX('{ABA_GASTOS}'!$B${LINHA_TOTAL_GASTOS}:"
                    f"$M${LINHA_TOTAL_GASTOS},1,{i})"
                ),
                o_que="gastos fixos do mes",
            )
        )
        lista.append(
            Regra(
                aba=ABA_CONTROLE,
                celula=f"F{linha}",
                formula=f"{ABA}!$B{linha}",
                o_que="fatura do cartao",
            )
        )

    # Linha TOTAL da aba Gastos Fixos: e a fonte de que o INDEX acima depende.
    for i in range(12):
        coluna = chr(ord("B") + i)
        lista.append(
            Regra(
                aba=ABA_GASTOS,
                celula=f"{coluna}{LINHA_TOTAL_GASTOS}",
                formula=f"SUM({coluna}4:{coluna}11)",
                o_que="total do mes",
            )
        )

    for linha in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1):
        lista.append(
            Regra(
                aba=ABA,
                celula=f"F{linha}",
                formula=formula_nao_classificado(linha).lstrip("="),
                o_que="nao classificado",
            )
        )

    lista.append(
        Regra(
            aba=ABA,
            celula=f"{COLUNA_GASTO}{ULTIMA_LINHA + 1}",
            formula=formula_total(COLUNA_GASTO, PRIMEIRA_LINHA, ULTIMA_LINHA),
            o_que="total do gasto do periodo",
        )
    )
    return lista


@dataclass(frozen=True)
class ValorEsperado:
    """O numero que a celula deve exibir, calculado a partir da fonte."""

    celula: str
    mes: str
    valor: float
    fonte: str


def valores_esperados(planilha: str | Path) -> list[ValorEsperado]:
    """Calcula o que as celulas ligadas devem mostrar quando a planilha abrir.

    Existe porque conferir o texto da formula nao basta: a formula pode estar
    certa no arquivo e a tela mostrar outra coisa - Excel aberto desde antes de
    a ferramenta escrever exibe a copia velha que ele tem em memoria, e nao o
    que esta em disco. Com os numeros na mao da para comparar com a tela e
    saber na hora se e isso.
    """
    import openpyxl

    wb = openpyxl.load_workbook(planilha)
    gastos = wb[ABA_GASTOS]
    cartao = wb[ABA]
    meses = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago",
             "set", "out", "nov", "dez"]

    esperados: list[ValorEsperado] = []
    for i, linha in enumerate(range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1)):
        coluna_mes = 2 + i  # B = janeiro
        total = sum(
            gastos.cell(r, coluna_mes).value or 0
            for r in range(4, 12)
            if isinstance(gastos.cell(r, coluna_mes).value, (int, float))
        )
        esperados.append(
            ValorEsperado(
                celula=f"E{linha}",
                mes=f"{meses[i]}/26",
                valor=float(total),
                fonte=f"soma de {ABA_GASTOS}!{chr(65 + coluna_mes - 1)}4:"
                f"{chr(65 + coluna_mes - 1)}11",
            )
        )

    for i, linha in enumerate(range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1)):
        valor = cartao[f"B{linha}"].value
        if isinstance(valor, (int, float)):
            esperados.append(
                ValorEsperado(
                    celula=f"F{linha}",
                    mes=f"{meses[i]}/26",
                    valor=float(valor),
                    fonte=f"{ABA}!B{linha}",
                )
            )
    wb.close()
    return esperados


def _normaliza(formula: str | None) -> str | None:
    """Compara so o essencial: o Excel varia aspas e espaco em branco."""
    if formula is None:
        return None
    return formula.replace(" ", "").replace("&quot;", '"').lstrip("=")


def confere(planilha: str | Path) -> list[Problema]:
    """Le a planilha e devolve as formulas quebradas. Nao escreve nada."""
    doc = Planilha(planilha)
    problemas: list[Problema] = []
    for regra in regras():
        aba = doc.aba(regra.aba)
        atual = aba.formula(regra.celula)
        if _normaliza(atual) != _normaliza(regra.formula):
            problemas.append(Problema(regra=regra, encontrado=atual))
    return problemas


def repara(planilha: str | Path, problemas: list[Problema]) -> int:
    """Reescreve as formulas quebradas. O backup e responsabilidade de quem chama.

    O estilo da celula e herdado de uma irma saudavel da mesma coluna quando a
    quebrada ficou com o estilo de texto que veio junto com o valor colado.
    """
    doc = Planilha(planilha)
    for problema in problemas:
        aba = doc.aba(problema.regra.aba)
        aba.define_formula(
            problema.regra.celula,
            problema.regra.formula,
            estilo=_estilo_saudavel(aba, problema.regra),
        )
    doc.recalcular_ao_abrir()
    doc.salva()
    return len(problemas)


def _estilo_saudavel(aba, regra: Regra) -> str | None:
    """O estilo da propria celula, a nao ser que ela tenha virado texto.

    Quando a formula e sobrescrita por um texto colado, o estilo do texto vem
    junto. Nesse caso vale mais o estilo de uma vizinha que escapou.
    """
    proprio = aba.estilo(regra.celula)
    achado = aba._localiza(regra.celula)
    virou_texto = bool(achado and 't="s"' in achado[2])
    if not virou_texto:
        return proprio

    coluna = regra.celula.rstrip("0123456789")
    for linha in range(PRIMEIRA_LINHA, ULTIMA_LINHA + 1):
        vizinha = f"{coluna}{linha}"
        if vizinha == regra.celula:
            continue
        if aba.formula(vizinha) is not None:
            return aba.estilo(vizinha)
    return proprio
