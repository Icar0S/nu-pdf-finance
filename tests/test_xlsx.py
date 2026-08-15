"""Testes da escrita cirurgica.

Estes testes olham o XML cru de proposito. A versao anterior escrevia com
openpyxl e conferia com openpyxl, o que passava verde enquanto o Excel abria a
planilha com a coluna de mes mostrando 46023 em vez de jan/26: openpyxl le o
numFmtId direto e ignora o `applyNumberFormat` que ele mesmo tinha perdido.

Conferir a saida de uma biblioteca com ela mesma nao prova nada sobre o que o
Excel vai mostrar.
"""

from __future__ import annotations

import re
import zipfile

import pytest

from nubank.errors import ErroExport
from nubank.xlsx import Aba, Planilha, indice_coluna

ABA_MINIMA = (
    '<?xml version="1.0" encoding="UTF-8"?>'
    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    "<cols>"
    '<col customWidth="1" min="1" max="1" width="11.0"/>'
    '<col customWidth="1" min="8" max="26" width="8.71"/>'
    "</cols>"
    "<sheetData>"
    '<row r="3"><c r="A3" s="31" t="s"><v>11</v></c>'
    '<c r="E3" s="31" t="s"><v>84</v></c></row>'
    '<row r="4"><c r="A4" s="10"><v>46023.0</v></c>'
    '<c r="C4" s="8"/><c r="E4" s="8"/>'
    '<c r="F4" s="19"><f t="shared" ref="F4:F15" si="1">'
    "IF($B4=&quot;&quot;,&quot;&quot;,$B4-SUM($C4:$E4))</f><v>1878.68</v></c>"
    "</row>"
    '<row r="16"><c r="A16" s="25" t="s"><v>64</v></c>'
    '<c r="E16" s="26"><f t="shared" si="2"/><v>0</v></c></row>'
    "</sheetData></worksheet>"
)


@pytest.fixture
def aba():
    return Aba(ABA_MINIMA, "Cartao")


@pytest.mark.parametrize("letra, esperado", [("A", 1), ("H", 8), ("Z", 26), ("AA", 27)])
def test_indice_coluna(letra, esperado):
    assert indice_coluna(letra) == esperado


def test_nao_confunde_h1_com_h16(aba):
    """Busca por prefixo casaria 'H1' dentro de 'H16'."""
    aba.define_numero("H16", 10.0, estilo="26")
    assert not aba.existe("H1")
    assert aba.existe("H16")


def test_preserva_o_estilo_da_celula_existente(aba):
    aba.define_numero("C4", 714.91)
    assert '<c r="C4" s="8"><v>714.91</v></c>' in aba.xml


def test_celula_nova_entra_em_ordem_de_coluna(aba):
    """O Excel exige as celulas ordenadas por coluna dentro da <row>."""
    aba.define_numero("D4", 1.0, estilo="8")
    row = re.search(r'<row r="4".*?</row>', aba.xml, re.DOTALL).group(0)
    refs = re.findall(r'<c r="([A-Z]+)\d+"', row)
    assert refs == sorted(refs, key=indice_coluna)
    assert "D" in refs


def test_formula_nao_leva_igual(aba):
    """Com '=' dentro de <f>, o Excel acusa arquivo corrompido."""
    aba.define_formula("F4", '=IF($B4="","",$H4-SUM($C4:$E4))')
    conteudo = re.search(r"<f[^>]*>(.*?)</f>", aba.xml, re.DOTALL).group(1)
    assert not conteudo.startswith("=")
    assert conteudo.startswith("IF(")


def test_formula_compartilhada_mantem_os_atributos(aba):
    """F4 e a mestra de F4:F15; perder ref/si quebraria as 11 irmas."""
    aba.define_formula("F4", '=IF($B4="","",$H4-SUM($C4:$E4))')
    assert 't="shared"' in aba.xml
    assert 'ref="F4:F15"' in aba.xml
    assert 'si="1"' in aba.xml


def test_formula_reescrita_perde_o_valor_em_cache(aba):
    """Senao o Excel mostra o numero velho ate alguem editar a planilha."""
    assert "<v>1878.68</v>" in aba.xml
    aba.define_formula("F4", '=IF($B4="","",$H4-SUM($C4:$E4))')
    celula = re.search(r'<c r="F4".*?</c>', aba.xml, re.DOTALL).group(0)
    assert "<v>" not in celula


def test_texto_vira_string_inline(aba):
    """Inline dispensa mexer no sharedStrings.xml, que e compartilhado."""
    aba.define_texto("G4", "29/11-30/12 | 85 lanc.", estilo="34")
    assert 't="inlineStr"' in aba.xml
    assert "<t xml:space=\"preserve\">29/11-30/12 | 85 lanc.</t>" in aba.xml


def test_texto_com_caractere_especial_e_escapado(aba):
    aba.define_texto("G4", "a < b & c", estilo="34")
    assert "a &lt; b &amp; c" in aba.xml


def test_largura_parte_o_intervalo_sobreposto(aba):
    """Dois <col> cobrindo a mesma coluna e arquivo invalido."""
    aba.largura_coluna("H", 18)

    cols = re.search(r"<cols>.*?</cols>", aba.xml, re.DOTALL).group(0)
    faixas = [
        (int(mn), int(mx))
        for mn, mx in re.findall(r'min="(\d+)"[^>]*max="(\d+)"', cols)
    ]
    assert (8, 8) in faixas
    assert (9, 26) in faixas
    assert (8, 26) not in faixas

    cobertas = [c for faixa in faixas for c in range(faixa[0], faixa[1] + 1)]
    assert len(cobertas) == len(set(cobertas)), "intervalos de coluna se sobrepoem"


def test_largura_nao_encolhe_coluna_ja_mais_larga(aba):
    aba.largura_coluna("A", 5)
    assert 'width="11.0"' in aba.xml


def test_linha_inexistente_da_erro_claro(aba):
    with pytest.raises(ErroExport, match="linha 99 nao existe"):
        aba.define_numero("H99", 1.0)


# --------------------------------------------------------------------------- #
# Planilha (zip)
# --------------------------------------------------------------------------- #


def test_so_reescreve_as_partes_alteradas(tmp_path):
    """A garantia central: o que nao foi pedido sai byte a byte igual."""
    import openpyxl

    origem = tmp_path / "p.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Cartao"
    wb.active["A1"] = 1
    wb.create_sheet("Outra")["A1"] = "intocada"
    wb.save(origem)

    antes = {n: zipfile.ZipFile(origem).read(n) for n in zipfile.ZipFile(origem).namelist()}

    doc = Planilha(origem)
    doc.aba("Cartao").define_numero("A1", 2.0)
    doc.salva()

    with zipfile.ZipFile(origem) as z:
        depois = {n: z.read(n) for n in z.namelist()}

    assert set(antes) == set(depois)
    alteradas = [n for n in antes if antes[n] != depois[n]]
    assert alteradas == ["xl/worksheets/sheet1.xml"]


def test_recalcular_ao_abrir_e_idempotente(tmp_path):
    import openpyxl

    origem = tmp_path / "p.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Cartao"
    wb.save(origem)

    doc = Planilha(origem)
    doc.recalcular_ao_abrir()
    doc.recalcular_ao_abrir()
    wbxml = doc.partes["xl/workbook.xml"].decode()
    assert wbxml.count("fullCalcOnLoad") == 1


def test_aba_inexistente_da_erro_claro(tmp_path):
    import openpyxl

    origem = tmp_path / "p.xlsx"
    openpyxl.Workbook().save(origem)
    with pytest.raises(ErroExport, match="nao tem a aba"):
        Planilha(origem).aba("Cartao")
