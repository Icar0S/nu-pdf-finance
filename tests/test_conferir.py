"""Testes do `conferir`.

O caso real que motivou tudo: formulas da coluna E do Controle Mensal foram
substituidas pelo texto 'Aporte mensal planejado' quando a planilha foi salva
por um aplicativo externo. Primeiro seis, depois nove. O Excel nao acusa erro
nenhum quando isso acontece - a conta so passa a dar outro numero.
"""

from __future__ import annotations

import shutil
import zipfile

import openpyxl
import pytest

from nubank.conferir import confere, regras, repara
from nubank.xlsx import Planilha

from .conftest import RAIZ

PLANILHA_REAL = RAIZ / "planejamento-avancado-financeiro-icaro26.xlsx"
requer_planilha = pytest.mark.skipif(
    not PLANILHA_REAL.exists(), reason="planilha real ausente"
)


@pytest.fixture
def planilha(tmp_path):
    copia = tmp_path / PLANILHA_REAL.name
    shutil.copy2(PLANILHA_REAL, copia)
    return copia


def test_regras_cobrem_as_ligacoes_entre_abas():
    por_aba = {}
    for r in regras():
        por_aba.setdefault(r.aba, []).append(r)
    assert set(por_aba) == {"Controle Mensal", "Gastos Fixos", "Cartao"}
    # 12 meses de gastos fixos + 12 de fatura do cartao
    assert len(por_aba["Controle Mensal"]) == 24
    assert len(por_aba["Gastos Fixos"]) == 12


@requer_planilha
def test_planilha_saudavel_nao_acusa_nada(planilha):
    assert confere(planilha) == []


@requer_planilha
def test_detecta_formula_virada_texto(planilha):
    """Exatamente o estrago observado: a celula vira um rotulo."""
    doc = Planilha(planilha)
    doc.aba("Controle Mensal").define_texto("E6", "Aporte mensal planejado")
    doc.salva()

    problemas = confere(planilha)
    assert len(problemas) == 1
    assert problemas[0].regra.celula == "E6"
    assert problemas[0].sumiu


@requer_planilha
def test_detecta_varias_e_repara_todas(planilha):
    doc = Planilha(planilha)
    aba = doc.aba("Controle Mensal")
    for linha in (4, 5, 6, 7, 11, 12, 13, 14, 15):
        aba.define_texto(f"E{linha}", "Aporte mensal planejado")
    doc.salva()

    problemas = confere(planilha)
    assert len(problemas) == 9

    assert repara(planilha, problemas) == 9
    assert confere(planilha) == []


@requer_planilha
def test_reparo_nao_deixa_a_celula_como_texto(planilha):
    """Restaurar a formula sem tirar o t="s" deixaria o Excel exibindo texto."""
    doc = Planilha(planilha)
    doc.aba("Controle Mensal").define_texto("E6", "Aporte mensal planejado")
    doc.salva()

    repara(planilha, confere(planilha))

    ws = openpyxl.load_workbook(planilha)["Controle Mensal"]
    assert str(ws["E6"].value).startswith("=INDEX(")


@requer_planilha
def test_reparo_so_toca_as_abas_com_problema(planilha):
    doc = Planilha(planilha)
    doc.aba("Controle Mensal").define_texto("E6", "Aporte mensal planejado")
    doc.salva()

    with zipfile.ZipFile(planilha) as z:
        antes = {n: z.read(n) for n in z.namelist()}

    repara(planilha, confere(planilha))

    with zipfile.ZipFile(planilha) as z:
        depois = {n: z.read(n) for n in z.namelist()}

    alteradas = {n for n in antes if antes[n] != depois[n]}
    assert alteradas <= {"xl/workbook.xml", "xl/worksheets/sheet2.xml"}
    assert antes["xl/styles.xml"] == depois["xl/styles.xml"]


@requer_planilha
def test_detecta_total_de_gastos_fixos_quebrado(planilha):
    """O INDEX aponta para essa linha; sem ela, tudo vira zero em silencio."""
    doc = Planilha(planilha)
    doc.aba("Gastos Fixos").define_numero("D13", 0.0)
    doc.salva()

    problemas = confere(planilha)
    assert [p.regra.celula for p in problemas] == ["D13"]


@requer_planilha
def test_reparo_e_idempotente(planilha):
    doc = Planilha(planilha)
    doc.aba("Controle Mensal").define_texto("E6", "Aporte mensal planejado")
    doc.salva()

    repara(planilha, confere(planilha))
    assert confere(planilha) == []
    assert repara(planilha, confere(planilha)) == 0
    assert confere(planilha) == []


@requer_planilha
def test_valores_esperados_batem_com_a_fonte(planilha):
    """O relatorio de valores existe para comparar com a tela do Excel.

    Se ele mentir, o usuario conclui que a planilha esta certa quando nao esta,
    ou o contrario. Confere contra a soma real da aba Gastos Fixos.
    """
    from nubank.conferir import valores_esperados

    ws = openpyxl.load_workbook(planilha)["Gastos Fixos"]
    esperados = {v.celula: v.valor for v in valores_esperados(planilha)}

    for i in range(12):
        total = sum(ws.cell(r, 2 + i).value or 0 for r in range(4, 12))
        assert esperados[f"E{4 + i}"] == pytest.approx(total)


@requer_planilha
def test_valores_esperados_cobrem_os_doze_meses(planilha):
    from nubank.conferir import valores_esperados

    celulas = {v.celula for v in valores_esperados(planilha)}
    assert {f"E{4 + i}" for i in range(12)} <= celulas
