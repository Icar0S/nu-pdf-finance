from datetime import datetime
from decimal import Decimal

import openpyxl
import pytest

from nubank.classify import carrega_tabela, classifica
from nubank.errors import ErroExport
from nubank.export import (
    aplica,
    exporta_csv,
    gasto_do_periodo,
    observacao,
    planeja,
)
from nubank.models import Bucket
from nubank.normalize import normaliza

from .conftest import RAIZ

PLANILHA_REAL = RAIZ / "planejamento-avancado-financeiro-icaro26.xlsx"
ARQUIVO_CARTAO = "sheet5.xml"  # a aba Cartao e a 5a da planilha


@pytest.fixture
def planilha(tmp_path):
    """Aba Cartao como ela nasceu: sem a coluna H e com a formula antiga de F.

    E de proposito que a fixture use a formula velha (conferindo contra B): e
    ela que o export precisa migrar.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartao"
    ws["A3"] = "Mes"
    ws["B3"] = "Fatura total"
    ws["C3"] = "Parcelas e assinaturas"
    ws["D3"] = "Essencial (mercado, farmacia, transporte)"
    ws["E3"] = "Superfluo (delivery, compras, saidas)"
    ws["F3"] = "Nao classificado"
    ws["G3"] = "Observacoes"
    for i, mes in enumerate(range(1, 13)):
        linha = 4 + i
        ws[f"A{linha}"] = datetime(2026, mes, 1)
        ws[f"F{linha}"] = f'=IF($B{linha}="","",$B{linha}-SUM($C{linha}:$E{linha}))'
    ws["A16"] = "TOTAL"
    for col in "BCDEF":
        ws[f"{col}16"] = f"=SUM({col}4:{col}15)"
    wb.create_sheet("Outra Aba")["A1"] = "nao pode ser tocada"
    caminho = tmp_path / "planilha.xlsx"
    wb.save(caminho)
    return caminho


@pytest.fixture
def fatura(raw):
    return classifica(normaliza(raw), carrega_tabela())


def test_dry_run_nao_escreve(planilha, fatura):
    antes = planilha.read_bytes()
    mudancas = planeja(planilha, [fatura])
    assert planilha.read_bytes() == antes
    assert any(m.mudou for m in mudancas)


def test_escreve_na_linha_certa(planilha, fatura):
    aplica(planilha, [fatura])
    ws = openpyxl.load_workbook(planilha)["Cartao"]
    linha = 4 + 7  # competencia 2026-08 -> agosto
    assert ws[f"B{linha}"].value == pytest.approx(float(fatura.total_a_pagar))
    assert ws[f"C{linha}"].value == pytest.approx(
        float(fatura.total_bucket(Bucket.PARCELAS))
    )
    assert ws["B4"].value is None  # janeiro fica intacto


def test_coluna_f_passa_a_conferir_contra_h(planilha, fatura):
    """A formula antiga conferia contra B e por isso nunca zerava."""
    ws = openpyxl.load_workbook(planilha)["Cartao"]
    assert "$B11-SUM(" in ws["F11"].value

    aplica(planilha, [fatura])

    ws = openpyxl.load_workbook(planilha)["Cartao"]
    assert ws["F11"].value == '=IF($B11="","",$H11-SUM($C11:$E11))'
    assert str(ws["F4"].value).startswith("=IF(")  # meses vazios tambem migram


def test_coluna_h_recebe_o_gasto_do_periodo(planilha, fatura):
    aplica(planilha, [fatura])
    ws = openpyxl.load_workbook(planilha)["Cartao"]
    assert ws["H3"].value == "Gasto do periodo"
    assert ws["H11"].value == pytest.approx(float(gasto_do_periodo(fatura)))
    assert ws["H16"].value == "=SUM(H4:H15)"


def test_c_mais_d_mais_e_somam_h_exatamente(planilha, fatura):
    """O ponto da coluna H: F volta a ser 'quanto falta classificar'.

    Exato, nao aproximado - senao F fica oscilando em 1 centavo por causa do
    arredondamento do resumo do PDF, e o zero deixa de ser legivel.
    """
    assert not fatura.pendentes
    aplica(planilha, [fatura])
    ws = openpyxl.load_workbook(planilha)["Cartao"]
    soma = sum(Decimal(str(ws[f"{c}11"].value)) for c in "CDE")
    assert soma == Decimal(str(ws["H11"].value))


def test_pendente_aparece_na_coluna_f(planilha, fatura):
    """Com --ignorar-pendentes, F mostra exatamente o que falta classificar."""
    pendente = fatura.transacoes[1]
    pendente.bucket = None  # simula merchant sem categoria
    aplica(planilha, [fatura])

    ws = openpyxl.load_workbook(planilha)["Cartao"]
    soma = sum(Decimal(str(ws[f"{c}11"].value)) for c in "CDE")
    assert Decimal(str(ws["H11"].value)) - soma == pendente.valor


def test_h_nasce_com_formato_de_moeda(planilha, fatura):
    """Sem copiar o estilo da vizinha, H sairia como numero cru."""
    aplica(planilha, [fatura])
    ws = openpyxl.load_workbook(planilha)["Cartao"]
    assert ws["H11"].number_format == ws["E11"].number_format
    assert ws.column_dimensions["H"].width >= 18


def test_backup_e_criado_antes_de_escrever(planilha, fatura):
    original = planilha.read_bytes()
    backup, _ = aplica(planilha, [fatura])
    assert backup.exists()
    assert backup.read_bytes() == original
    assert planilha.read_bytes() != original


def test_outras_abas_sobrevivem(planilha, fatura):
    aplica(planilha, [fatura])
    wb = openpyxl.load_workbook(planilha)
    assert wb["Outra Aba"]["A1"].value == "nao pode ser tocada"


def test_observacao_explica_a_diferenca_entre_b_e_h(fatura):
    """B e H divergem pelo saldo carregado; isso precisa estar escrito."""
    texto = observacao(fatura)
    assert "B-H" in texto
    assert "saldo da fatura anterior" in texto


def test_segunda_aplicacao_nao_muda_nada(planilha, fatura):
    aplica(planilha, [fatura])
    assert not [m for m in planeja(planilha, [fatura]) if m.mudou]


def test_planilha_sem_a_aba_cartao(tmp_path, fatura):
    caminho = tmp_path / "vazia.xlsx"
    openpyxl.Workbook().save(caminho)
    with pytest.raises(ErroExport, match="aba 'Cartao'"):
        planeja(caminho, [fatura])


def test_cabecalho_inesperado(planilha, fatura):
    wb = openpyxl.load_workbook(planilha)
    wb["Cartao"]["B3"] = "Outra coisa"
    wb.save(planilha)
    with pytest.raises(ErroExport, match="formato esperado"):
        planeja(planilha, [fatura])


def test_competencia_sem_linha_na_planilha(planilha, fatura):
    wb = openpyxl.load_workbook(planilha)
    wb["Cartao"]["A11"] = datetime(2027, 8, 1)
    wb.save(planilha)
    with pytest.raises(ErroExport, match="2026-08"):
        planeja(planilha, [fatura])


def test_csv_de_detalhe(tmp_path, fatura):
    destino = exporta_csv(tmp_path / "detalhe.csv", [fatura])
    linhas = destino.read_text(encoding="utf-8-sig").splitlines()
    assert linhas[0].startswith("competencia;data;descricao")
    assert len(linhas) == 1 + len(fatura.transacoes)


@pytest.mark.skipif(not PLANILHA_REAL.exists(), reason="planilha real ausente")
def test_escrita_na_planilha_real_so_toca_a_aba_cartao(tmp_path, fatura):
    """A garantia forte: nenhuma outra parte do arquivo e reescrita.

    Comparado no zip cru, e nao pelo openpyxl. Foi conferir openpyxl com
    openpyxl que deixou passar a perda dos atributos apply* dos estilos, que no
    Excel virava a coluna de mes mostrando 46023 em vez de jan/26.
    """
    import shutil
    import zipfile

    copia = tmp_path / PLANILHA_REAL.name
    shutil.copy2(PLANILHA_REAL, copia)

    with zipfile.ZipFile(copia) as z:
        antes = {n: z.read(n) for n in z.namelist()}

    aplica(copia, [fatura])

    with zipfile.ZipFile(copia) as z:
        depois = {n: z.read(n) for n in z.namelist()}

    assert set(antes) == set(depois), "alguma parte do arquivo sumiu ou apareceu"
    alteradas = {n for n in antes if antes[n] != depois[n]}

    # workbook.xml so entra quando o fullCalcOnLoad ainda nao estava marcado,
    # entao ele e opcional; a aba Cartao e obrigatoria.
    permitidas = {"xl/workbook.xml", f"xl/worksheets/{ARQUIVO_CARTAO}"}
    assert alteradas <= permitidas, f"partes inesperadas: {alteradas - permitidas}"
    assert f"xl/worksheets/{ARQUIVO_CARTAO}" in alteradas

    # As que mais importam: estilos e strings sao compartilhados por todas as
    # abas, e um deslize neles estraga a planilha inteira.
    for parte in ("xl/styles.xml", "xl/sharedStrings.xml", "xl/worksheets/sheet2.xml"):
        assert antes[parte] == depois[parte], f"{parte} foi reescrito"


@pytest.mark.skipif(not PLANILHA_REAL.exists(), reason="planilha real ausente")
def test_escrita_na_planilha_real_preserva_as_flags_de_estilo(tmp_path, fatura):
    """Sem applyNumberFormat="1", o Excel ignora o numFmtId e usa General."""
    import shutil
    import zipfile

    copia = tmp_path / PLANILHA_REAL.name
    shutil.copy2(PLANILHA_REAL, copia)

    def flags(caminho):
        with zipfile.ZipFile(caminho) as z:
            estilos = z.read("xl/styles.xml").decode()
        return {
            f: estilos.count(f'{f}="1"')
            for f in ("applyNumberFormat", "applyFont", "applyBorder", "applyFill")
        }

    antes = flags(copia)
    aplica(copia, [fatura])
    assert flags(copia) == antes
    assert antes["applyNumberFormat"] > 0, "fixture nao exercita o caso"
